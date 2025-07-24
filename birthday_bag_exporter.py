import pandas as pd
import re
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import subprocess
import threading
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

# Check and install required packages
def install_requirements():
    try:
        # Check if tkinterdnd2 is installed
        import importlib.util
        tkdnd_spec = importlib.util.find_spec("tkinterdnd2")
        
        if tkdnd_spec is None:
            print("Installing required packages...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-r", "requirements.txt"])
            print("Packages installed successfully!")
            messagebox.showinfo("Setup Complete", "Required packages have been installed. Please restart the application.")
            sys.exit(0)
    except Exception as e:
        print(f"Error installing packages: {str(e)}")
        messagebox.showerror("Installation Error", f"Error installing required packages: {str(e)}\n\nPlease run 'pip install -r requirements.txt' manually.")

# Try to import tkinterdnd2
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except ImportError:
    # If not available, install requirements and notify user to restart
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    install_requirements()
    root.destroy()
    # Continue with basic tkinter for now
    TkinterDnD = tk.Tk

class BirthdayBagExporter:
    def __init__(self, root):
        self.root = root
        self.root.title("Birthday Bag Exporter")
        self.root.geometry("900x650")
        
        # Try to set the application icon
        try:
            if os.path.exists("icon.ico"):
                self.root.iconbitmap("icon.ico")
            elif os.path.exists("icon.png"):
                icon = tk.PhotoImage(file="icon.png")
                self.root.iconphoto(True, icon)
        except Exception as e:
            print(f"Could not load icon: {e}")
        
        # Theme settings
        self.dark_mode = tk.BooleanVar(value=False)
        self.apply_theme()
        
        # Route assignments dictionary - easily editable
        self.route_assignments = {
            # SUN routes
            'SUN_CANYON COUNTRY-2': '8',
            'SUN_NORTHEAST-2': '2',
            'SUN_NORTHEAST-3': '3',
            'SUN_SOUTH GATE-5': '4',
            'SUN_WESTMONT PARK-1': '5',
            'SUN_WESTSIDE-1': '6',
            'SUN_EAST LA-4': '7',
            'SUN_CANYON CNTRY-2': '8',
            'SUN_ROSEGATE-1': '9',
            'SUN_CULVER CITY': '10',
            'SUN_VERDUGO': '11',
            
            # MON routes
            'MON_CANYON CNTRY-1': '1',
            'MON_EAST LA-1': '2',
            'MON_EAST LA-2': '3',
            'MON_EAST LA-3': '4',
            'MON_MID CITY-1': '5',
            'MON_MID CITY-2': '6',
            'MON_MID CITY-3': '7',
            'MON_ROSECRANS-1': '8',
            'MON_SOUTH GATE-1': '9',
            'MON_SAN GABRIEL-2': '10',
            'MON_ROSECRANS-2': '11',
            'MON_MID CITY-4': 'VOL-1',
            'MON_CANYON CNTRY-3': 'VOL-2',
            
            # TUE routes
            'TUE_NO. HOLLYWOOD-2': '1',
            'TUE_NORTH VALLEY-1': '2',
            'TUE_NORTH VALLEY-2': '3',
            'TUE_SILVERLAKE-1': '4',
            'TUE_SOUTH GATE-2': '5',
            'TUE_SOUTH GATE-3': '6',
            'TUE_VERNON-1': '7',
            'TUE_WAC-1': '8',
            'TUE_WAC-4': '9',
            'TUE_VERNON-2': '10',
            'TUE_SOUTH LA-1': '11',
            'TUE_NOHO-1': 'VOLUNTEER-1',
            'TUE_BRAD': 'VOLUNTEER-2',
            'TUE_THE ELITE': 'VOLUNTEER-3',
            'TUE_PICK UP LIST': 'VOLUNTEER-4',
            
            # WED routes
            'WED_SOUTH BAY-4': '1',
            'WED_LAUREL-2': '2',
            'WED_LONG BEACH-1': '3',
            'WED_SAN GABRIEL-1': '4',
            'WED_SEPULVEDA-2': '5',
            'WED_SILVERLAKE-3': '6',
            'WED_SOUTH BAY-1': '7',
            'WED_SOUTHEAST-2': '8',
            'WED_WAC-2': '9',
            'WED_WRHAP-1': '10',
            'WED_SOUTHEAST-4': '11',
            'WED_LAUREL-1': 'VOLUNTEER-1',
            'WED_COA NEIGHBORHOOD': 'VOLUNTEER-2',
            
            # THU routes
            'THU_AGAPE-1': '1',
            'THU_DOWNTOWN-1': '2',
            'THU_HOLLYWOOD-1': '3',
            'THU_LA PUENTE-1': '4',
            'THU_LANCASTER-1': '5',
            'THU_LANCASTER-2': '6',
            'THU_SEPULVEDA-1': '7',
            'THU_SOUTH BAY-2': '8',
            'THU_SOUTHEAST-1': '9',
            'THU_WESTSIDE-2': '10',
            'THU_SOUTH BAY PLUS': '11',
            'THU_HOLLYWOOD-2': 'VOLUNTEER-1',
            'THU_PALM PLUS': 'VOLUNTEER-2',
            
            # FRI routes
            'FRI_HOLLYWOOD-3': '1',
            'FRI_HUNTINGTON PARK-1': '2',
            'FRI_NORTHEAST-1': '3',
            'FRI_SOUTH BAY-3': '4',
            'FRI_SOUTH GATE-4': '5',
            'FRI_WAC-3': '6',
            'FRI_WEST VALLEY-1': '7',
            'FRI_WEST VALLEY-2': '8',
            'FRI_WEST VALLEY-3': '9',
            'FRI_SOUTHEAST-3': '10',
            'FRI_HUNTINGTON PARK-2': '11',
        }
        
        # Create UI elements
        self.create_widgets()
        
        # Set up drag and drop if available
        try:
            self.root.drop_target_register(DND_FILES)
            self.root.dnd_bind('<<Drop>>', self.drop)
            
            # Add drag-drop indicator
            self.drop_label.config(text="Drag and drop your Excel file here\nor click Browse to select a file")
        except Exception:
            # Drag and drop not available
            self.drop_label.config(text="Click Browse to select your Excel file")
    
    def apply_theme(self):
        """Apply light or dark theme based on current setting"""
        style = ttk.Style()
        
        if self.dark_mode.get():
            # Dark mode
            self.root.configure(bg="#2d2d2d")
            style.configure("TFrame", background="#2d2d2d")
            style.configure("TLabel", background="#2d2d2d", foreground="#ffffff")
            style.configure("TButton", background="#444444", foreground="#ffffff")
            style.configure("Accent.TButton", background="#007acc", foreground="#ffffff")
            style.configure("TCheckbutton", background="#2d2d2d", foreground="#ffffff")
            style.configure("TEntry", fieldbackground="#3d3d3d", foreground="#ffffff")
            style.map("TCheckbutton", background=[("active", "#3d3d3d")])
            style.map("TButton", background=[("active", "#555555")])
            style.map("Accent.TButton", background=[("active", "#0088cc")])
            
            # Configure the drop zone
            style.configure("Drop.TFrame", background="#3d3d3d", bordercolor="#555555")
        else:
            # Light mode
            self.root.configure(bg="#f5f5f5")
            style.configure("TFrame", background="#f5f5f5")
            style.configure("TLabel", background="#f5f5f5", foreground="#000000")
            style.configure("TButton", background="#e1e1e1", foreground="#000000")
            style.configure("Accent.TButton", background="#007acc", foreground="#ffffff")
            style.configure("TCheckbutton", background="#f5f5f5", foreground="#000000")
            style.configure("TEntry", fieldbackground="#ffffff", foreground="#000000")
            style.map("TCheckbutton", background=[("active", "#e5e5e5")])
            style.map("TButton", background=[("active", "#d0d0d0")])
            style.map("Accent.TButton", background=[("active", "#0088cc")])
            
            # Configure the drop zone
            style.configure("Drop.TFrame", background="#ffffff", bordercolor="#cccccc")
        
        # Update all widgets with the new theme
        self.update_all_widgets()
    
    def update_all_widgets(self):
        """Update all widgets with the current theme"""
        # This will be called after theme changes
        if hasattr(self, 'drop_frame'):
            if self.dark_mode.get():
                self.drop_frame.configure(style="Drop.TFrame")
                self.drop_label.configure(foreground="#ffffff")
            else:
                self.drop_frame.configure(style="Drop.TFrame")
                self.drop_label.configure(foreground="#000000")
    
    def toggle_theme(self):
        """Toggle between light and dark mode"""
        self.dark_mode.set(not self.dark_mode.get())
        self.apply_theme()
    
    def create_widgets(self):
        # Create a main frame with padding
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create a header frame
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 20))
        
        # App title
        title_label = ttk.Label(
            header_frame, 
            text="Birthday Bag Exporter",
            font=("Arial", 18, "bold")
        )
        title_label.pack(side=tk.LEFT)
        
        # Dark mode toggle
        theme_frame = ttk.Frame(header_frame)
        theme_frame.pack(side=tk.RIGHT)
        
        theme_check = ttk.Checkbutton(
            theme_frame,
            text="Dark Mode",
            variable=self.dark_mode,
            command=self.toggle_theme
        )
        theme_check.pack(side=tk.RIGHT)
        
        # Create a content frame with a nice border
        content_frame = ttk.Frame(main_frame, padding="15")
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Drop zone frame
        self.drop_frame = ttk.Frame(content_frame, padding="20", relief="solid", borderwidth=1, style="Drop.TFrame")
        self.drop_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Drop zone label
        self.drop_label = ttk.Label(
            self.drop_frame, 
            text="Loading...",
            font=("Arial", 12)
        )
        self.drop_label.pack(pady=30)
        
        # File path display
        self.file_path_var = tk.StringVar()
        file_path_frame = ttk.Frame(content_frame)
        file_path_frame.pack(fill=tk.X, pady=10)
        
        file_path_label = ttk.Label(file_path_frame, text="Selected File:", font=("Arial", 10, "bold"))
        file_path_label.pack(side=tk.LEFT, padx=5)
        
        file_path_entry = ttk.Entry(file_path_frame, textvariable=self.file_path_var, width=50)
        file_path_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        browse_button = ttk.Button(file_path_frame, text="Browse", command=self.browse_file)
        browse_button.pack(side=tk.LEFT, padx=5)
        
        # Output file options
        output_frame = ttk.Frame(content_frame)
        output_frame.pack(fill=tk.X, pady=10)
        
        output_label = ttk.Label(output_frame, text="Output File:", font=("Arial", 10, "bold"))
        output_label.pack(side=tk.LEFT, padx=5)
        
        self.output_var = tk.StringVar(value="Birthday_Bag_Routes.xlsx")
        output_entry = ttk.Entry(output_frame, textvariable=self.output_var, width=50)
        output_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # Button frame
        button_frame = ttk.Frame(content_frame)
        button_frame.pack(fill=tk.X, pady=20)
        
        # Process button
        process_button = ttk.Button(
            button_frame, 
            text="Process File", 
            command=self.process_file,
            style="Accent.TButton"
        )
        process_button.pack(side=tk.LEFT, padx=5)
        
        # Route assignments editor button
        edit_button = ttk.Button(
            button_frame, 
            text="Edit Route Assignments", 
            command=self.open_route_editor
        )
        edit_button.pack(side=tk.LEFT, padx=5)
        
        # Status frame
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill=tk.X, pady=10)
        
        # Status label
        self.status_var = tk.StringVar(value="Ready")
        status_label = ttk.Label(
            status_frame, 
            textvariable=self.status_var,
            font=("Arial", 10)
        )
        status_label.pack(side=tk.LEFT, padx=5)
        
        # Progress bar
        self.progress = ttk.Progressbar(status_frame, orient="horizontal", length=300, mode="determinate")
        self.progress.pack(side=tk.RIGHT, padx=5)
        
        # Create a custom style for the accent button
        style = ttk.Style()
        style.configure("Accent.TButton", font=("Arial", 11, "bold"))
    
    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Birthday Labels File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            self.file_path_var.set(file_path)
    
    def drop(self, event):
        file_path = event.data
        # Clean up the file path (remove curly braces and quotes if present)
        file_path = file_path.strip('{}')
        if file_path.startswith('"') and file_path.endswith('"'):
            file_path = file_path[1:-1]
        
        if file_path.lower().endswith(('.xlsx', '.xls')):
            self.file_path_var.set(file_path)
        else:
            messagebox.showerror("Invalid File", "Please drop an Excel file (.xlsx or .xls)")
    
    def process_file(self):
        input_file = self.file_path_var.get()
        output_file = self.output_var.get()
        
        if not input_file:
            messagebox.showerror("Error", "Please select an input file")
            return
        
        if not output_file:
            messagebox.showerror("Error", "Please specify an output file name")
            return
        
        # Start processing in a separate thread to keep UI responsive
        threading.Thread(target=self.process_file_thread, args=(input_file, output_file), daemon=True).start()
    
    def process_file_thread(self, input_file, output_file):
        try:
            self.status_var.set("Processing...")
            self.progress["value"] = 0
            self.root.update_idletasks()
            
            # Process the file
            self.process_route_data(input_file, output_file)
            
            self.status_var.set(f"Done! Output saved to {output_file}")
            self.progress["value"] = 100
            self.root.update_idletasks()
            
            # Show success message in the main thread
            self.root.after(0, lambda: messagebox.showinfo("Success", f"File processed successfully!\nOutput saved to {output_file}"))
        except Exception as e:
            err_msg = str(e)
            self.status_var.set(f"Error: {err_msg}")
            self.root.after(0, lambda: messagebox.showerror("Error", f"An error occurred: {err_msg}"))
    
    def open_route_editor(self):
        # Create a new window for editing route assignments
        editor_window = tk.Toplevel(self.root)
        editor_window.title("Route Assignments Editor")
        editor_window.geometry("900x700")
        
        # Apply the current theme to the editor window
        if self.dark_mode.get():
            editor_window.configure(bg="#2d2d2d")
        else:
            editor_window.configure(bg="#f5f5f5")
        
        # Create a frame for the editor
        editor_frame = ttk.Frame(editor_window, padding="20")
        editor_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header
        header_label = ttk.Label(
            editor_frame, 
            text="Route Assignments Editor",
            font=("Arial", 16, "bold")
        )
        header_label.pack(pady=(0, 20))
        
        # Create a notebook for tabs (one tab per day)
        notebook = ttk.Notebook(editor_frame)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # Create tabs for each day
        days = ['SUN', 'MON', 'TUE', 'WED', 'THU', 'FRI']
        day_frames = {}
        day_data = {}
        
        for day in days:
            # Create a frame for this day
            day_frames[day] = ttk.Frame(notebook, padding="15")
            notebook.add(day_frames[day], text=day)
            
            # Create a scrollable frame
            canvas = tk.Canvas(day_frames[day], highlightthickness=0)
            if self.dark_mode.get():
                canvas.configure(bg="#2d2d2d")
            else:
                canvas.configure(bg="#f5f5f5")
                
            scrollbar = ttk.Scrollbar(day_frames[day], orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e, canvas=canvas: canvas.configure(
                    scrollregion=canvas.bbox("all")
                )
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Create headers
            ttk.Label(scrollable_frame, text="Van Number", font=("Arial", 11, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky="w")
            ttk.Label(scrollable_frame, text="Route Name", font=("Arial", 11, "bold")).grid(row=0, column=1, padx=5, pady=5, sticky="w")
            
            # Get routes for this day
            day_routes = {k.split('_', 1)[1]: v for k, v in self.route_assignments.items() if k.startswith(f"{day}_")}
            day_data[day] = {}
            
            # Sort routes by van number
            sorted_routes = sorted(day_routes.items(), key=lambda x: self.sort_key(x[1]))
            
            # Create entry fields for each route
            row = 1
            for route, van in sorted_routes:
                van_var = tk.StringVar(value=van)
                day_data[day][route] = van_var
                
                ttk.Entry(scrollable_frame, textvariable=van_var, width=10).grid(row=row, column=0, padx=5, pady=2)
                ttk.Label(scrollable_frame, text=route).grid(row=row, column=1, padx=5, pady=2, sticky="w")
                row += 1
            
            # Add button to add new route
            add_frame = ttk.Frame(scrollable_frame)
            add_frame.grid(row=row, column=0, columnspan=2, pady=20, sticky="w")
            
            ttk.Label(add_frame, text="Add New Route:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
            
            new_van_var = tk.StringVar()
            ttk.Label(add_frame, text="Van #:").pack(side=tk.LEFT, padx=5)
            ttk.Entry(add_frame, textvariable=new_van_var, width=10).pack(side=tk.LEFT, padx=5)
            
            new_route_var = tk.StringVar()
            ttk.Label(add_frame, text="Route Name:").pack(side=tk.LEFT, padx=5)
            ttk.Entry(add_frame, textvariable=new_route_var, width=20).pack(side=tk.LEFT, padx=5)
            
            ttk.Button(
                add_frame, 
                text="Add", 
                command=lambda d=day, r=new_route_var, v=new_van_var, sf=scrollable_frame, dd=day_data: self.add_route(d, r, v, sf, dd)
            ).pack(side=tk.LEFT, padx=5)
        
        # Button frame
        button_frame = ttk.Frame(editor_frame)
        button_frame.pack(pady=20)
        
        # Add save button
        save_button = ttk.Button(
            button_frame, 
            text="Save Changes", 
            command=lambda: self.save_route_assignments(day_data, editor_window),
            style="Accent.TButton"
        )
        save_button.pack(side=tk.LEFT, padx=5)
        
        # Add cancel button
        cancel_button = ttk.Button(
            button_frame, 
            text="Cancel", 
            command=editor_window.destroy
        )
        cancel_button.pack(side=tk.LEFT, padx=5)
    
    def sort_key(self, van):
        """Helper function to sort van numbers correctly"""
        if van.isdigit():
            return int(van)
        elif van.startswith('VOLUNTEER-') and van[10:].isdigit():
            return 100 + int(van[10:])  # Put VOLUNTEER-# after numeric vans
        elif van.startswith('VOL-') and van[4:].isdigit():
            return 100 + int(van[4:])  # Put VOL-# after numeric vans
        else:
            # For other non-numeric values, put them at the end
            return float('inf')
    
    def add_route(self, day, route_var, van_var, scrollable_frame, day_data):
        route = route_var.get().strip()
        van = van_var.get().strip()
        
        if not route or not van:
            messagebox.showerror("Error", "Please enter both route name and van number")
            return
        
        # Add to the UI - need to recreate the entire list to maintain sorting
        day_data[day][route] = tk.StringVar(value=van)
        
        # Clear all existing route entries
        for widget in scrollable_frame.winfo_children():
            if isinstance(widget, ttk.Frame):  # Keep the add frame
                continue
            widget.destroy()
        
        # Recreate headers
        ttk.Label(scrollable_frame, text="Van Number", font=("Arial", 11, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Label(scrollable_frame, text="Route Name", font=("Arial", 11, "bold")).grid(row=0, column=1, padx=5, pady=5, sticky="w")
        
        # Sort routes by van number
        sorted_routes = sorted(day_data[day].items(), key=lambda x: self.sort_key(x[1].get()))
        
        # Recreate all route entries
        row = 1
        for route_name, existing_van_var in sorted_routes:
            ttk.Entry(scrollable_frame, textvariable=existing_van_var, width=10).grid(row=row, column=0, padx=5, pady=2)
            ttk.Label(scrollable_frame, text=route_name).grid(row=row, column=1, padx=5, pady=2, sticky="w")
            row += 1

        # Clear entry fields
        route_var.set("")
        van_var.set("")
    
    def save_route_assignments(self, day_data, editor_window):
        # Update the route assignments dictionary
        new_assignments = {}
        
        for day, routes in day_data.items():
            for route, van_var in routes.items():
                key = f"{day}_{route}"
                new_assignments[key] = van_var.get()
        
        self.route_assignments = new_assignments
        messagebox.showinfo("Success", "Route assignments saved successfully!")
        editor_window.destroy()
    
    def extract_client_data(self, report_file):
        """
        Extract client data from Sheet1 of the Report file
        This function will extract client names and their routes
        """
        self.update_progress(10, "Extracting client data...")
        
        # Read the Excel file
        df = pd.read_excel(report_file, sheet_name='Sheet1', header=None)
        
        # Initialize lists to store extracted data
        clients = []
        routes = []
        
        # Pattern to match route information (day and route name)
        route_pattern = re.compile(r'(MON|TUE|WED|THU|FRI|SAT|SUN)\s+\|\s+(.*?)$')
        
        # Iterate through the dataframe to extract client and route information
        i = 0
        while i < len(df):
            # Check if this row contains "Happy Birthday!"
            if isinstance(df.iloc[i, 0], str) and "Happy Birthday!" in df.iloc[i, 0]:
                # Next row should contain client name
                if i + 1 < len(df):
                    client_name = df.iloc[i + 1, 0]
                    
                    # Look for route information in nearby rows (up to 5 rows after client name)
                    route_found = False
                    for j in range(i + 2, min(i + 7, len(df))):
                        for col in range(df.shape[1]):
                            if pd.notna(df.iloc[j, col]) and isinstance(df.iloc[j, col], str):
                                match = route_pattern.search(df.iloc[j, col])
                                if match and not route_found:
                                    day = match.group(1)
                                    route_name = match.group(2)
                                    full_route = f"{day} | {route_name}"
                                    clients.append(client_name)
                                    routes.append(full_route)
                                    route_found = True
                                    break
                        if route_found:
                            break
            i += 1
        
        # Create a dataframe with the extracted data
        client_data = pd.DataFrame({
            'Client': clients,
            'Route': routes
        })
        
        return client_data
    
    def match_clients_to_vans(self, client_data):
        """
        Match clients to van numbers based on their routes
        """
        self.update_progress(30, "Matching clients to vans...")
        
        # Initialize list to store van numbers
        van_numbers = []
        
        # Process each client
        for _, row in client_data.iterrows():
            route = row['Route']
            # Extract day and route name
            match = re.match(r'(MON|TUE|WED|THU|FRI|SAT|SUN)\s+\|\s+(.*?)$', route)
            if match:
                day = match.group(1)
                route_name = match.group(2).strip()
                
                # Clean route name for matching
                cleaned_route = self.clean_route_name(route_name)
                
                # Try to find a match in route assignments
                van_num = self.find_van_for_route(day, cleaned_route)
                
                van_numbers.append(van_num if van_num else "")
            else:
                van_numbers.append("")
        
        # Add van numbers to client data
        client_data['VAN #'] = van_numbers
        
        return client_data
    
    def clean_route_name(self, route_name):
        """
        Clean and standardize route names for better matching
        """
        # Remove any text in parentheses
        cleaned = re.sub(r'\s*\(.*?\)', '', route_name)
        # Remove [FULL] tag
        cleaned = re.sub(r'\[FULL\]\s*', '', cleaned)
        # Remove extra spaces
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()
        # Convert to uppercase for case-insensitive comparison
        return cleaned.upper()
    
    def find_van_for_route(self, day, route_name):
        """
        Find the van number for a given day and route name
        """
        # Try exact match first
        key = f"{day}_{route_name}"
        if key in self.route_assignments:
            return self.route_assignments[key]
        
        # Try to match the base route name
        base_route = self.extract_route_base(route_name)
        for k, v in self.route_assignments.items():
            if k.startswith(f"{day}_") and base_route in k:
                return v
        
        # Try partial matching
        for k, v in self.route_assignments.items():
            if k.startswith(f"{day}_"):
                route_part = k.split('_', 1)[1]
                if base_route in route_part or route_part in base_route:
                    return v
        
        return ""
    
    def extract_route_base(self, route_name):
        """
        Extract the base part of the route name (e.g., "SOUTH BAY" from "SOUTH BAY-1")
        """
        # Extract the part before any dash or number
        match = re.match(r'([A-Za-z\s]+)', route_name)
        if match:
            return match.group(1).strip()
        return route_name
    
    def order_by_day_and_van(self, client_data):
        """
        Order the client data by day of week and then by van number
        """
        self.update_progress(50, "Ordering by day and van...")
        
        # Define day order
        day_order = {
            'SUN': 0,
            'MON': 1,
            'TUE': 2,
            'WED': 3,
            'THU': 4,
            'FRI': 5,
            'SAT': 6
        }
        
        # Extract day from route
        client_data['Day'] = client_data['Route'].apply(
            lambda x: re.match(r'(MON|TUE|WED|THU|FRI|SAT|SUN)', x).group(1) if re.match(r'(MON|TUE|WED|THU|FRI|SAT|SUN)', x) else ""
        )
        
        # Convert van numbers to integers for sorting (if possible)
        client_data['VAN_SORT'] = client_data['VAN #'].apply(
            lambda x: int(x) if isinstance(x, str) and x.isdigit() else float('inf')
        )
        
        # Map days to their order
        client_data['DAY_SORT'] = client_data['Day'].map(day_order)
        
        # Sort by day and then by van number
        client_data = client_data.sort_values(['DAY_SORT', 'VAN_SORT'])
        
        # Drop temporary columns
        client_data = client_data.drop(['DAY_SORT', 'VAN_SORT'], axis=1)
        
        return client_data
    
    def format_output(self, client_data):
        """
        Format the output to match Sheet2 format
        """
        self.update_progress(60, "Formatting output...")
        
        # Reorder columns
        formatted_data = client_data[['VAN #', 'Client', 'Route', 'Day']]
        
        # Rename columns
        formatted_data = formatted_data.rename(columns={'Route': 'Route Name & Day'})
        
        # Add Notes column
        formatted_data['Notes'] = ""
        
        return formatted_data
    
    def add_day_separators(self, formatted_data):
        """
        Add empty rows between different days to match the Sheet2 format
        """
        self.update_progress(70, "Adding day separators...")
        
        # Get unique days in order
        days = formatted_data['Day'].unique()
        
        # Create a new dataframe with separators
        result = []
        
        for day in days:
            # Get rows for this day
            day_rows = formatted_data[formatted_data['Day'] == day]
            
            # Add rows to result
            for _, row in day_rows.iterrows():
                result.append(row)
            
            # Add separator row (except after the last day)
            if day != days[-1]:
                # Add a single black bar row with the day value for identification
                separator_row = pd.Series({'VAN #': '', 'Client': '', 'Route Name & Day': '', 'Notes': '', 'Day': f"SEPARATOR_{day}"})
                result.append(separator_row)
        
        # Convert result to dataframe
        result_df = pd.DataFrame(result)
        
        return result_df
    
    def apply_excel_formatting(self, output_file):
        """
        Apply formatting to the Excel file to match the example with merged cells for black bars
        """
        self.update_progress(80, "Applying Excel formatting...")
        
        # Load the workbook
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Define styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        
        # Apply formatting to header row
        for col in range(1, 5):  # A to D (VAN #, Client, Route Name & Day, Notes)
            cell = ws.cell(row=1, column=col)
            cell.border = thin_border
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Apply borders to all cells and identify separator rows
        max_row = ws.max_row
        separator_rows = []
        
        for row in range(1, max_row + 1):
            # Check if this is a separator row
            day_cell = ws.cell(row=row, column=5)  # Column E contains the Day
            
            if day_cell.value and isinstance(day_cell.value, str) and day_cell.value.startswith("SEPARATOR_"):
                separator_rows.append(row)
            else:
                # Apply borders to regular rows
                for col in range(1, 5):  # A to D
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
        
        # Process separator rows - merge cells and apply black fill
        for row in separator_rows:
            # Merge cells A to D
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            
            # Apply black fill to the merged cell
            merged_cell = ws.cell(row=row, column=1)
            merged_cell.fill = black_fill
            
            # Remove any text
            merged_cell.value = ""
            
            # Remove borders from the merged cell
            merged_cell.border = None
        
        # Hide the Day column (E)
        ws.column_dimensions['E'].hidden = True
        
        # Set column widths to match the screenshot
        ws.column_dimensions['A'].width = 10  # VAN #
        ws.column_dimensions['B'].width = 30  # Client
        ws.column_dimensions['C'].width = 40  # Route Name & Day
        ws.column_dimensions['D'].width = 15  # Notes
        
        # Add a black bar at the top
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        top_bar = ws.cell(row=1, column=1)
        top_bar.fill = black_fill
        top_bar.border = None

        # Save the workbook
        wb.save(output_file)
        self.update_progress(90, "Finalizing...")

    def update_progress(self, value, message=None):
        """Update progress bar and status message"""
        self.progress["value"] = value
        if message:
            self.status_var.set(message)
        self.root.update_idletasks()
    def process_route_data(self, report_file, output_file):
        """
        Main function to process route data
        """
        self.update_progress(0, "Starting...")
        
        client_data = self.extract_client_data(report_file)
        client_data = self.match_clients_to_vans(client_data)
        client_data = self.order_by_day_and_van(client_data)
        formatted_data = self.format_output(client_data)
        final_data = self.add_day_separators(formatted_data)
        self.update_progress(75, f"Saving to {output_file}...")
        final_data.to_excel(output_file, sheet_name='Sheet1', index=False)

        self.apply_excel_formatting(output_file)

        self.update_progress(100, f"Done! Output saved to {output_file}")
        return final_data


def main():
    try:
        # Try to use TkinterDnD for drag and drop
        root = TkinterDnD.Tk()
    except Exception:
        # Fall back to regular Tk
        root = tk.Tk()

    BirthdayBagExporter(root)
    root.mainloop()


if __name__ == "__main__":
    main()
